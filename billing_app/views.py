from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum, Count, Q
from django.utils import timezone
from datetime import datetime, timedelta, date
from decimal import Decimal
import json
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .models import Menu, Order, OrderItem, Expense
from .forms import MenuForm, ExpenseForm
from .utils import (
    generate_qr_code_response,
    calculate_daily_sales,
    calculate_monthly_sales,
    calculate_expenses,
    calculate_profit,
    calculate_yearly_sales
)


# Admin Authentication Views
def admin_login(request):
    """Admin login page"""
    if request.user.is_authenticated:
        return redirect('admin_dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            return redirect('admin_dashboard')
        else:
            messages.error(request, 'Invalid username or password')
    
    return render(request, 'admin/login.html')


@login_required
def admin_logout(request):
    """Admin logout - redirects to login page"""
    logout(request)
    messages.success(request, 'You have been logged out successfully.')
    return redirect('admin_login')


@login_required
def admin_dashboard(request):
    """Admin dashboard - shows only data created after user account creation"""
    today = timezone.now().date()
    user_created_date = request.user.date_joined.date()
    
    # Filter data created on or after user account creation date
    # This ensures new superusers only see their own data, not old data
    filter_date = max(user_created_date, today)
    
    # Today's statistics (only orders created after user account)
    today_orders = Order.objects.filter(
        created_at__date=today,
        created_at__gte=request.user.date_joined
    )
    today_sales = sum(order.total_amount for order in today_orders.filter(status='paid'))
    today_total_orders = today_orders.count()
    
    # Total menu items (only items created after user account)
    total_menu_items = Menu.objects.filter(created_at__gte=request.user.date_joined).count()
    available_items = Menu.objects.filter(
        is_available=True,
        created_at__gte=request.user.date_joined
    ).count()
    
    # Check if user is new (created today)
    is_new_user = user_created_date == today
    
    context = {
        'today_sales': today_sales,
        'today_total_orders': today_total_orders,
        'total_menu_items': total_menu_items,
        'available_items': available_items,
        'is_new_user': is_new_user,
        'user_created_date': user_created_date,
    }
    
    return render(request, 'admin/dashboard.html', context)


# Menu Management Views
@login_required
def menu_list(request):
    """List all menu items - only shows items created after user account"""
    menu_items = Menu.objects.filter(created_at__gte=request.user.date_joined)
    return render(request, 'admin/menu_list.html', {'menu_items': menu_items})


@login_required
def menu_add(request):
    """Add new menu item"""
    if request.method == 'POST':
        form = MenuForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Menu item added successfully!')
            return redirect('menu_list')
    else:
        form = MenuForm()
    
    return render(request, 'admin/menu_form.html', {'form': form, 'action': 'Add'})


@login_required
def menu_edit(request, pk):
    """Edit menu item"""
    menu_item = get_object_or_404(Menu, pk=pk)
    
    if request.method == 'POST':
        form = MenuForm(request.POST, instance=menu_item)
        if form.is_valid():
            form.save()
            messages.success(request, 'Menu item updated successfully!')
            return redirect('menu_list')
    else:
        form = MenuForm(instance=menu_item)
    
    return render(request, 'admin/menu_form.html', {'form': form, 'menu_item': menu_item, 'action': 'Edit'})


@login_required
def menu_delete(request, pk):
    """Delete menu item"""
    menu_item = get_object_or_404(Menu, pk=pk)
    
    if request.method == 'POST':
        menu_item.delete()
        messages.success(request, 'Menu item deleted successfully!')
        return redirect('menu_list')
    
    return render(request, 'admin/menu_delete.html', {'menu_item': menu_item})


@login_required
def menu_toggle_availability(request, pk):
    """Toggle menu item availability"""
    menu_item = get_object_or_404(Menu, pk=pk)
    menu_item.is_available = not menu_item.is_available
    menu_item.save()
    
    status = 'available' if menu_item.is_available else 'unavailable'
    messages.success(request, f'Menu item marked as {status}!')
    return redirect('menu_list')


# Reports Views
@login_required
def reports_dashboard(request):
    """Reports dashboard - shows only data created after user account"""
    today = timezone.now().date()
    current_month = timezone.now().month
    current_year = timezone.now().year
    user_date_joined = request.user.date_joined
    
    # Daily sales (filtered by user)
    daily_sales = calculate_daily_sales(today, user_date_joined)
    
    # Monthly sales (filtered by user)
    monthly_sales = calculate_monthly_sales(current_year, current_month, user_date_joined)
    
    # Current month expenses (filtered by user)
    month_start = datetime(current_year, current_month, 1).date()
    month_expenses = calculate_expenses(month_start, today, user_date_joined)
    
    # Profit calculation
    profit = calculate_profit(monthly_sales['total_sales'], month_expenses['total_expenses'])
    
    # Yearly sales (from user creation date to today)
    yearly_start_date = max(user_date_joined.date(), date(current_year, 1, 1))
    yearly_sales = calculate_yearly_sales(yearly_start_date, today, user_date_joined)
    
    context = {
        'daily_sales': daily_sales,
        'monthly_sales': monthly_sales,
        'expenses': month_expenses,
        'profit': profit,
        'yearly_sales': yearly_sales,
        'today': today,
        'current_month': current_month,
        'current_year': current_year,
    }
    
    return render(request, 'admin/reports.html', context)


# Expenses Management Views
@login_required
def expense_list(request):
    """List and filter expenses - only shows expenses created after user account"""
    expenses = Expense.objects.filter(created_at__gte=request.user.date_joined)
    return render(request, 'admin/expense_list.html', {'expenses': expenses})


@login_required
def expense_add(request):
    """Add new expense"""
    if request.method == 'POST':
        form = ExpenseForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Expense added successfully!')
            return redirect('expense_list')
    else:
        form = ExpenseForm()
    return render(request, 'admin/expense_form.html', {'form': form, 'action': 'Add'})


@login_required
def expense_edit(request, pk):
    """Edit expense"""
    expense = get_object_or_404(Expense, pk=pk)
    if request.method == 'POST':
        form = ExpenseForm(request.POST, instance=expense)
        if form.is_valid():
            form.save()
            messages.success(request, 'Expense updated successfully!')
            return redirect('expense_list')
    else:
        form = ExpenseForm(instance=expense)
    return render(request, 'admin/expense_form.html', {'form': form, 'action': 'Edit', 'expense': expense})


@login_required
def expense_delete(request, pk):
    """Delete expense"""
    expense = get_object_or_404(Expense, pk=pk)
    if request.method == 'POST':
        expense.delete()
        messages.success(request, 'Expense deleted successfully!')
        return redirect('expense_list')
    return render(request, 'admin/expense_delete.html', {'expense': expense})


# Billing Interface Views
def billing_index(request):
    """Public billing interface"""
    menu_items = Menu.objects.filter(is_available=True)
    return render(request, 'billing/index.html', {'menu_items': menu_items})


def add_to_cart(request):
    """Add item to cart via AJAX"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            menu_item_id = data.get('menu_item_id')
            
            menu_item = get_object_or_404(Menu, pk=menu_item_id, is_available=True)
            
            # Get or create cart from session
            cart = request.session.get('cart', {})
            item_key = str(menu_item_id)
            
            if item_key in cart:
                cart[item_key]['quantity'] += 1
            else:
                cart[item_key] = {
                    'id': menu_item.id,
                    'name': menu_item.name,
                    'price': str(menu_item.price),
                    'quantity': 1,
                }
            
            request.session['cart'] = cart
            request.session.modified = True
            
            return JsonResponse({
                'success': True,
                'message': f'{menu_item.name} added to cart',
                'cart_count': sum(item['quantity'] for item in cart.values())
            })
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': 'Invalid request'})


def get_cart(request):
    """Get cart contents via AJAX"""
    cart = request.session.get('cart', {})
    total = sum(Decimal(item['price']) * item['quantity'] for item in cart.values())
    
    return JsonResponse({
        'cart': list(cart.values()),
        'total': str(total),
        'cart_count': sum(item['quantity'] for item in cart.values())
    })


def update_cart_item(request):
    """Update cart item quantity via AJAX"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            menu_item_id = str(data.get('menu_item_id'))
            quantity = int(data.get('quantity', 1))
            
            cart = request.session.get('cart', {})
            
            if menu_item_id in cart:
                if quantity <= 0:
                    del cart[menu_item_id]
                else:
                    cart[menu_item_id]['quantity'] = quantity
                
                request.session['cart'] = cart
                request.session.modified = True
                
                total = sum(Decimal(item['price']) * item['quantity'] for item in cart.values())
                
                return JsonResponse({
                    'success': True,
                    'cart_count': sum(item['quantity'] for item in cart.values()),
                    'total': str(total)
                })
            else:
                return JsonResponse({'success': False, 'message': 'Item not found in cart'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': 'Invalid request'})


def remove_from_cart(request):
    """Remove item from cart via AJAX"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            menu_item_id = str(data.get('menu_item_id'))
            
            cart = request.session.get('cart', {})
            
            if menu_item_id in cart:
                del cart[menu_item_id]
                request.session['cart'] = cart
                request.session.modified = True
                
                total = sum(Decimal(item['price']) * item['quantity'] for item in cart.values())
                
                return JsonResponse({
                    'success': True,
                    'cart_count': sum(item['quantity'] for item in cart.values()),
                    'total': str(total)
                })
            else:
                return JsonResponse({'success': False, 'message': 'Item not found in cart'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': 'Invalid request'})


def clear_cart(request):
    """Clear entire cart"""
    request.session['cart'] = {}
    request.session.modified = True
    return JsonResponse({'success': True, 'message': 'Cart cleared'})


def create_order(request):
    """Create order from cart"""
    if request.method == 'POST':
        try:
            cart = request.session.get('cart', {})
            
            if not cart:
                return JsonResponse({'success': False, 'message': 'Cart is empty'})
            
            # Calculate total
            total_amount = sum(Decimal(item['price']) * item['quantity'] for item in cart.values())
            
            # Create order
            order = Order.objects.create(total_amount=total_amount, status='pending')
            
            # Create order items
            for item in cart.values():
                menu_item = Menu.objects.get(pk=item['id'])
                OrderItem.objects.create(
                    order=order,
                    menu_item=menu_item,
                    quantity=item['quantity'],
                    price=Decimal(item['price'])
                )
            
            # Clear cart
            request.session['cart'] = {}
            request.session.modified = True
            
            return JsonResponse({
                'success': True,
                'order_id': order.id,
                'order_number': order.order_number,
                'message': 'Order created successfully!'
            })
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': 'Invalid request'})


def view_bill(request, order_id):
    """View bill for an order"""
    order = get_object_or_404(Order, pk=order_id)
    return render(request, 'billing/bill.html', {'order': order})


def pay_now(request, order_id):
    """Mark order as paid"""
    if request.method == 'POST':
        order = get_object_or_404(Order, pk=order_id)
        order.status = 'paid'
        order.save()
        return JsonResponse({'success': True, 'message': 'Order marked as paid!'})
    
    return JsonResponse({'success': False, 'message': 'Invalid request'})


def generate_qr(request, order_id):
    """Generate QR code for an order"""
    order = get_object_or_404(Order, pk=order_id)
    return generate_qr_code_response(order)


def export_bill_pdf(request, order_id):
    """Export bill as PDF - same format as print (NO QR CODE)"""
    order = get_object_or_404(Order, pk=order_id)
    
    response = HttpResponse(content_type='application/pdf')
    # Use bill1.pdf as filename if it's the first bill, otherwise use order number
    response['Content-Disposition'] = f'attachment; filename="bill1.pdf"'
    
    # Use letter size but can adapt to any size
    doc = SimpleDocTemplate(response, pagesize=letter, 
                           leftMargin=0.5*inch, rightMargin=0.5*inch,
                           topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Header
    header_style = styles['Heading1']
    header_style.alignment = 1  # Center alignment
    header_style.fontSize = 20
    header = Paragraph("Restaurant Bill", header_style)
    elements.append(header)
    elements.append(Spacer(1, 0.2*inch))
    
    # Order Number
    order_num_style = styles['Normal']
    order_num_style.alignment = 1
    order_num_style.fontSize = 14
    order_num = Paragraph(f"Order #: {order.order_number}", order_num_style)
    elements.append(order_num)
    elements.append(Spacer(1, 0.1*inch))
    
    # Date
    date_style = styles['Normal']
    date_style.alignment = 1
    date_style.fontSize = 10
    date_str = order.created_at.strftime('%B %d, %Y %I:%M %p')
    date_para = Paragraph(f"Date: {date_str}", date_style)
    elements.append(date_para)
    elements.append(Spacer(1, 0.3*inch))
    
    # Items Table
    table_data = [['Item', 'Qty', 'Price (₹)', 'Subtotal (₹)']]
    
    for item in order.order_items.all():
        table_data.append([
            item.menu_item.name,
            str(item.quantity),
            f"₹{item.price:.2f}",
            f"₹{item.subtotal:.2f}"
        ])
    
    # Total Row
    table_data.append([
        '',
        '',
        '<b>Total</b>',
        f"<b>₹{order.total_amount:.2f}</b>"
    ])
    
    # Create table - NO QR CODE included
    bill_table = Table(table_data, colWidths=[3*inch, 0.8*inch, 1.2*inch, 1*inch])
    bill_table.setStyle(TableStyle([
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
        ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        
        # Data rows
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 10),
        ('BOTTOMPADDING', (0, 1), (-1, -2), 8),
        ('TOPPADDING', (0, 1), (-1, -2), 8),
        ('GRID', (0, 0), (-1, -2), 1, colors.black),
        
        # Total row
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 11),
        ('TOPPADDING', (0, -1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 10),
        ('LINEABOVE', (0, -1), (-1, -1), 2, colors.black),
        ('GRID', (0, -1), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(bill_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Status
    status_style = styles['Normal']
    status_style.alignment = 1
    status_style.fontSize = 10
    status_text = f"Status: {order.status.upper()}"
    status_para = Paragraph(status_text, status_style)
    elements.append(status_para)
    
    # NOTE: QR CODE IS INTENTIONALLY NOT INCLUDED - Same as print format
    # Build PDF
    doc.build(elements)
    return response


# Export Views
@login_required
def export_daily_pdf(request, year, month, day):
    """Export daily sales report as PDF"""
    from datetime import date
    report_date = date(int(year), int(month), int(day))
    daily_sales = calculate_daily_sales(report_date, request.user.date_joined)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="daily_sales_{report_date}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title = Paragraph(f"Daily Sales Report - {report_date}", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))
    
    # Summary
    data = [
        ['Date', str(report_date)],
        ['Total Sales', f"₹{daily_sales['total_sales']:.2f}"],
        ['Total Orders', str(daily_sales['total_orders'])],
    ]
    
    table = Table(data, colWidths=[2*inch, 4*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.grey),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (1, 0), (1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 0.5*inch))
    
    # Orders table
    if daily_sales['orders']:
        elements.append(Paragraph("Orders", styles['Heading2']))
        order_data = [['Order Number', 'Amount', 'Status', 'Time']]
        for order in daily_sales['orders']:
            order_data.append([
                order.order_number,
                f"₹{order.total_amount:.2f}",
                order.status,
                order.created_at.strftime('%H:%M:%S')
            ])
        
        order_table = Table(order_data, colWidths=[2*inch, 1.5*inch, 1*inch, 1.5*inch])
        order_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(order_table)
    
    doc.build(elements)
    return response


@login_required
def export_daily_excel(request, year, month, day):
    """Export daily sales report as Excel"""
    report_date = date(int(year), int(month), int(day))
    daily_sales = calculate_daily_sales(report_date, request.user.date_joined)
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Daily Sales {report_date}"
    
    # Header
    ws['A1'] = f"Daily Sales Report - {report_date}"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:D1')
    
    # Summary
    ws['A3'] = 'Date'
    ws['B3'] = str(report_date)
    ws['A4'] = 'Total Sales'
    ws['B4'] = f"₹{daily_sales['total_sales']:.2f}"
    ws['A5'] = 'Total Orders'
    ws['B5'] = daily_sales['total_orders']
    
    # Orders table
    if daily_sales['orders']:
        ws['A7'] = 'Order Number'
        ws['B7'] = 'Amount'
        ws['C7'] = 'Status'
        ws['D7'] = 'Time'
        
        for i, order in enumerate(daily_sales['orders'], start=8):
            ws[f'A{i}'] = order.order_number
            ws[f'B{i}'] = f"${order.total_amount:.2f}"
            ws[f'C{i}'] = order.status
            ws[f'D{i}'] = order.created_at.strftime('%H:%M:%S')
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="daily_sales_{report_date}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_monthly_pdf(request, year, month):
    """Export monthly sales report as PDF"""
    monthly_sales = calculate_monthly_sales(int(year), int(month), request.user.date_joined)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="monthly_sales_{year}_{month}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title = Paragraph(f"Monthly Sales Report - {month}/{year}", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))
    
    # Summary
    data = [
        ['Month', f"{month}/{year}"],
        ['Total Sales', f"₹{monthly_sales['total_sales']:.2f}"],
        ['Total Orders', str(monthly_sales['total_orders'])],
    ]
    
    table = Table(data, colWidths=[2*inch, 4*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.grey),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (1, 0), (1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(table)
    
    doc.build(elements)
    return response


@login_required
def export_monthly_excel(request, year, month):
    """Export monthly sales report as Excel"""
    monthly_sales = calculate_monthly_sales(int(year), int(month), request.user.date_joined)
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Monthly Sales {month}/{year}"
    
    # Header
    ws['A1'] = f"Monthly Sales Report - {month}/{year}"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:D1')
    
    # Summary
    ws['A3'] = 'Month'
    ws['B3'] = f"{month}/{year}"
    ws['A4'] = 'Total Sales'
    ws['B4'] = f"₹{monthly_sales['total_sales']:.2f}"
    ws['A5'] = 'Total Orders'
    ws['B5'] = monthly_sales['total_orders']
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="monthly_sales_{year}_{month}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_expenses_pdf(request, start_date, end_date):
    """Export expenses report as PDF"""
    expenses_data = calculate_expenses(start_date, end_date, request.user.date_joined)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="expenses_{start_date}_{end_date}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title = Paragraph(f"Expenses Report - {start_date} to {end_date}", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))
    
    # Summary
    data = [
        ['Period', f"{start_date} to {end_date}"],
        ['Total Expenses', f"₹{expenses_data['total_expenses']:.2f}"],
    ]
    
    table = Table(data, colWidths=[2*inch, 4*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.grey),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (1, 0), (1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 0.5*inch))
    
    # Expenses table
    if expenses_data['expenses']:
        elements.append(Paragraph("Expenses", styles['Heading2']))
        expense_data = [['Date', 'Description', 'Category', 'Amount']]
        for expense in expenses_data['expenses']:
            expense_data.append([
                str(expense.date),
                expense.description,
                expense.get_category_display(),
                f"₹{expense.amount:.2f}"
            ])
        
        expense_table = Table(expense_data, colWidths=[1.5*inch, 2.5*inch, 1.5*inch, 1*inch])
        expense_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(expense_table)
    
    doc.build(elements)
    return response


@login_required
def export_expenses_excel(request, start_date, end_date):
    """Export expenses report as Excel"""
    expenses_data = calculate_expenses(start_date, end_date, request.user.date_joined)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"
    
    # Header
    ws['A1'] = f"Expenses Report - {start_date} to {end_date}"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:D1')
    
    # Summary
    ws['A3'] = 'Period'
    ws['B3'] = f"{start_date} to {end_date}"
    ws['A4'] = 'Total Expenses'
    ws['B4'] = f"₹{expenses_data['total_expenses']:.2f}"
    
    # Expenses table
    if expenses_data['expenses']:
        ws['A6'] = 'Date'
        ws['B6'] = 'Description'
        ws['C6'] = 'Category'
        ws['D6'] = 'Amount'
        
        for i, expense in enumerate(expenses_data['expenses'], start=7):
            ws[f'A{i}'] = str(expense.date)
            ws[f'B{i}'] = expense.description
            ws[f'C{i}'] = expense.get_category_display()
            ws[f'D{i}'] = f"₹{expense.amount:.2f}"
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="expenses_{start_date}_{end_date}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_profit_pdf(request, start_date, end_date):
    """Export profit report as PDF"""
    monthly_sales = calculate_monthly_sales(timezone.now().year, timezone.now().month, request.user.date_joined)
    expenses_data = calculate_expenses(start_date, end_date, request.user.date_joined)
    profit = calculate_profit(monthly_sales['total_sales'], expenses_data['total_expenses'])
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="profit_{start_date}_{end_date}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title = Paragraph(f"Profit Report - {start_date} to {end_date}", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))
    
    # Summary
    data = [
        ['Period', f"{start_date} to {end_date}"],
        ['Total Sales', f"₹{monthly_sales['total_sales']:.2f}"],
        ['Total Expenses', f"₹{expenses_data['total_expenses']:.2f}"],
        ['Profit', f"₹{profit:.2f}"],
    ]
    
    table = Table(data, colWidths=[2*inch, 4*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.grey),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (1, 0), (1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (1, 3), (1, 3), colors.lightgreen if profit > 0 else colors.lightcoral),
    ]))
    elements.append(table)
    
    doc.build(elements)
    return response


@login_required
def export_profit_excel(request, start_date, end_date):
    """Export profit report as Excel"""
    monthly_sales = calculate_monthly_sales(timezone.now().year, timezone.now().month, request.user.date_joined)
    expenses_data = calculate_expenses(start_date, end_date, request.user.date_joined)
    profit = calculate_profit(monthly_sales['total_sales'], expenses_data['total_expenses'])
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Profit"
    
    # Header
    ws['A1'] = f"Profit Report - {start_date} to {end_date}"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:B1')
    
    # Summary
    ws['A3'] = 'Period'
    ws['B3'] = f"{start_date} to {end_date}"
    ws['A4'] = 'Total Sales'
    ws['B4'] = f"₹{monthly_sales['total_sales']:.2f}"
    ws['A5'] = 'Total Expenses'
    ws['B5'] = f"₹{expenses_data['total_expenses']:.2f}"
    ws['A6'] = 'Profit'
    ws['B6'] = f"₹{profit:.2f}"
    ws['B6'].font = Font(size=14, bold=True, color='00FF00' if profit > 0 else 'FF0000')
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="profit_{start_date}_{end_date}.xlsx"'
    wb.save(response)
    return response


@login_required
def reports_graphs(request):
    """Reports graphs page with daily and monthly sales/items graphs"""
    today = timezone.now().date()
    current_month = timezone.now().month
    current_year = timezone.now().year
    user_date_joined = request.user.date_joined
    
    # Get last 30 days for daily graph
    daily_data = []
    daily_items_data = {}
    daily_breakdown = []  # Detailed daily breakdown
    
    for i in range(29, -1, -1):
        date = today - timedelta(days=i)
        daily_sales = calculate_daily_sales(date, user_date_joined)
        daily_data.append({
            'date': date.strftime('%Y-%m-%d'),
            'sales': float(daily_sales['total_sales']),
            'orders': daily_sales['total_orders']
        })
        
        # Daily breakdown data
        daily_breakdown.append({
            'date': date,
            'date_str': date.strftime('%Y-%m-%d'),
            'sales': float(daily_sales['total_sales']),
            'orders': daily_sales['total_orders']
        })
        
        # Get items for this day
        orders = Order.objects.filter(
            created_at__date=date,
            status='paid',
            created_at__gte=user_date_joined
        )
        for order in orders:
            for item in order.order_items.all():
                item_name = item.menu_item.name
                if item_name not in daily_items_data:
                    daily_items_data[item_name] = {}
                if date.strftime('%Y-%m-%d') not in daily_items_data[item_name]:
                    daily_items_data[item_name][date.strftime('%Y-%m-%d')] = 0
                daily_items_data[item_name][date.strftime('%Y-%m-%d')] += item.quantity
    
    # Top Items - Last 7 Days
    top_items_7days = {}
    for i in range(6, -1, -1):
        date = today - timedelta(days=i)
        orders = Order.objects.filter(
            created_at__date=date,
            status='paid',
            created_at__gte=user_date_joined
        )
        for order in orders:
            for item in order.order_items.all():
                item_name = item.menu_item.name
                if item_name not in top_items_7days:
                    top_items_7days[item_name] = {'quantity': 0, 'revenue': Decimal('0')}
                top_items_7days[item_name]['quantity'] += item.quantity
                top_items_7days[item_name]['revenue'] += item.price * item.quantity
    
    top_items_7days_list = sorted(
        [{'name': k, 'quantity': v['quantity'], 'revenue': float(v['revenue'])} 
         for k, v in top_items_7days.items()],
        key=lambda x: x['quantity'],
        reverse=True
    )
    
    # Top Items - Last 6 Months
    top_items_6months = {}
    for i in range(5, -1, -1):
        target_date = today - timedelta(days=30*i)
        year = target_date.year
        month = target_date.month
        orders = Order.objects.filter(
            created_at__year=year,
            created_at__month=month,
            status='paid',
            created_at__gte=user_date_joined
        )
        for order in orders:
            for item in order.order_items.all():
                item_name = item.menu_item.name
                if item_name not in top_items_6months:
                    top_items_6months[item_name] = {'quantity': 0, 'revenue': Decimal('0')}
                top_items_6months[item_name]['quantity'] += item.quantity
                top_items_6months[item_name]['revenue'] += item.price * item.quantity
    
    top_items_6months_list = sorted(
        [{'name': k, 'quantity': v['quantity'], 'revenue': float(v['revenue'])} 
         for k, v in top_items_6months.items()],
        key=lambda x: x['quantity'],
        reverse=True
    )
    
    # Get last 12 months for monthly graph
    monthly_data = []
    monthly_items_data = {}
    monthly_breakdown = []  # Detailed monthly breakdown
    
    for i in range(11, -1, -1):
        target_date = today - timedelta(days=30*i)
        year = target_date.year
        month = target_date.month
        monthly_sales = calculate_monthly_sales(year, month, user_date_joined)
        monthly_data.append({
            'month': f"{year}-{month:02d}",
            'sales': float(monthly_sales['total_sales']),
            'orders': monthly_sales['total_orders']
        })
        
        # Monthly breakdown data
        monthly_breakdown.append({
            'year': year,
            'month': month,
            'month_str': f"{year}-{month:02d}",
            'sales': float(monthly_sales['total_sales']),
            'orders': monthly_sales['total_orders']
        })
        
        # Get items for this month
        orders = Order.objects.filter(
            created_at__year=year,
            created_at__month=month,
            status='paid',
            created_at__gte=user_date_joined
        )
        for order in orders:
            for item in order.order_items.all():
                item_name = item.menu_item.name
                month_key = f"{year}-{month:02d}"
                if item_name not in monthly_items_data:
                    monthly_items_data[item_name] = {}
                if month_key not in monthly_items_data[item_name]:
                    monthly_items_data[item_name][month_key] = 0
                monthly_items_data[item_name][month_key] += item.quantity
    
    context = {
        'daily_data': json.dumps(daily_data),
        'daily_items_data': json.dumps(daily_items_data),
        'monthly_data': json.dumps(monthly_data),
        'monthly_items_data': json.dumps(monthly_items_data),
        'daily_breakdown': daily_breakdown,
        'top_items_7days': top_items_7days_list,
        'top_items_6months': top_items_6months_list,
        'monthly_breakdown': monthly_breakdown,
    }
    
    return render(request, 'admin/reports_graphs.html', context)


@login_required
def yearly_sales_report(request):
    """Yearly sales report with date range"""
    today = timezone.now().date()
    user_date_joined = request.user.date_joined
    
    # Get start date from request or use user creation date
    start_date_str = request.GET.get('start_date')
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except:
            start_date = max(user_date_joined.date(), date(today.year, 1, 1))
    else:
        start_date = max(user_date_joined.date(), date(today.year, 1, 1))
    
    end_date = today
    
    yearly_sales = calculate_yearly_sales(start_date, end_date, user_date_joined)
    
    context = {
        'yearly_sales': yearly_sales,
        'start_date': start_date,
        'end_date': end_date,
    }
    
    return render(request, 'admin/yearly_sales_report.html', context)


@login_required
def export_yearly_pdf(request):
    """Export yearly sales report as PDF"""
    today = timezone.now().date()
    user_date_joined = request.user.date_joined
    
    start_date_str = request.GET.get('start_date')
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except:
            start_date = max(user_date_joined.date(), date(today.year, 1, 1))
    else:
        start_date = max(user_date_joined.date(), date(today.year, 1, 1))
    
    end_date = today
    yearly_sales = calculate_yearly_sales(start_date, end_date, user_date_joined)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="yearly_sales_{start_date}_{end_date}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title = Paragraph(f"Yearly Sales Report - {start_date} to {end_date}", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))
    
    # Summary
    data = [
        ['Period', f"{start_date} to {end_date}"],
        ['Total Sales', f"₹{yearly_sales['total_sales']:.2f}"],
        ['Total Orders', str(yearly_sales['total_orders'])],
    ]
    
    table = Table(data, colWidths=[2*inch, 4*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.grey),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (1, 0), (1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 0.5*inch))
    
    # Item breakdown
    if yearly_sales['item_breakdown']:
        elements.append(Paragraph("Item-wise Sales", styles['Heading2']))
        item_data = [['Item Name', 'Quantity', 'Revenue (₹)']]
        for item in yearly_sales['item_breakdown']:
            item_data.append([
                item['name'],
                str(item['total_quantity']),
                f"₹{item['total_revenue']:.2f}"
            ])
        
        item_table = Table(item_data, colWidths=[3*inch, 1.5*inch, 1.5*inch])
        item_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(item_table)
    
    doc.build(elements)
    return response


@login_required
def export_yearly_excel(request):
    """Export yearly sales report as Excel"""
    today = timezone.now().date()
    user_date_joined = request.user.date_joined
    
    start_date_str = request.GET.get('start_date')
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except:
            start_date = max(user_date_joined.date(), date(today.year, 1, 1))
    else:
        start_date = max(user_date_joined.date(), date(today.year, 1, 1))
    
    end_date = today
    yearly_sales = calculate_yearly_sales(start_date, end_date, user_date_joined)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Yearly Sales"
    
    # Header
    ws['A1'] = f"Yearly Sales Report - {start_date} to {end_date}"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:C1')
    
    # Summary
    ws['A3'] = 'Period'
    ws['B3'] = f"{start_date} to {end_date}"
    ws['A4'] = 'Total Sales'
    ws['B4'] = f"₹{yearly_sales['total_sales']:.2f}"
    ws['A5'] = 'Total Orders'
    ws['B5'] = yearly_sales['total_orders']
    
    # Item breakdown
    if yearly_sales['item_breakdown']:
        ws['A7'] = 'Item Name'
        ws['B7'] = 'Quantity'
        ws['C7'] = 'Revenue (₹)'
        
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        for col in ['A7', 'B7', 'C7']:
            ws[col].fill = header_fill
            ws[col].font = header_font
        
        for i, item in enumerate(yearly_sales['item_breakdown'], start=8):
            ws[f'A{i}'] = item['name']
            ws[f'B{i}'] = item['total_quantity']
            ws[f'C{i}'] = f"₹{item['total_revenue']:.2f}"
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="yearly_sales_{start_date}_{end_date}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_daily_breakdown_pdf(request):
    """Export daily breakdown as PDF"""
    today = timezone.now().date()
    user_date_joined = request.user.date_joined
    
    daily_breakdown = []
    for i in range(29, -1, -1):
        date = today - timedelta(days=i)
        daily_sales = calculate_daily_sales(date, user_date_joined)
        daily_breakdown.append({
            'date': date,
            'sales': float(daily_sales['total_sales']),
            'orders': daily_sales['total_orders']
        })
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="daily_breakdown_{today}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    title = Paragraph("Daily Breakdown - Last 30 Days", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))
    
    data = [['Date', 'Sales (₹)', 'Orders']]
    for item in daily_breakdown:
        data.append([
            item['date'].strftime('%Y-%m-%d'),
            f"₹{item['sales']:.2f}",
            str(item['orders'])
        ])
    
    table = Table(data, colWidths=[2*inch, 2*inch, 2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(table)
    
    doc.build(elements)
    return response


@login_required
def export_daily_breakdown_excel(request):
    """Export daily breakdown as Excel"""
    today = timezone.now().date()
    user_date_joined = request.user.date_joined
    
    daily_breakdown = []
    for i in range(29, -1, -1):
        date = today - timedelta(days=i)
        daily_sales = calculate_daily_sales(date, user_date_joined)
        daily_breakdown.append({
            'date': date,
            'sales': float(daily_sales['total_sales']),
            'orders': daily_sales['total_orders']
        })
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Breakdown"
    
    ws['A1'] = "Daily Breakdown - Last 30 Days"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:C1')
    
    ws['A3'] = 'Date'
    ws['B3'] = 'Sales (₹)'
    ws['C3'] = 'Orders'
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for col in ['A3', 'B3', 'C3']:
        ws[col].fill = header_fill
        ws[col].font = header_font
    
    for i, item in enumerate(daily_breakdown, start=4):
        ws[f'A{i}'] = item['date'].strftime('%Y-%m-%d')
        ws[f'B{i}'] = f"₹{item['sales']:.2f}"
        ws[f'C{i}'] = item['orders']
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="daily_breakdown_{today}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_top_items_7days_pdf(request):
    """Export top items last 7 days as PDF"""
    today = timezone.now().date()
    user_date_joined = request.user.date_joined
    
    top_items = {}
    for i in range(6, -1, -1):
        date = today - timedelta(days=i)
        orders = Order.objects.filter(
            created_at__date=date,
            status='paid',
            created_at__gte=user_date_joined
        )
        for order in orders:
            for item in order.order_items.all():
                item_name = item.menu_item.name
                if item_name not in top_items:
                    top_items[item_name] = {'quantity': 0, 'revenue': Decimal('0')}
                top_items[item_name]['quantity'] += item.quantity
                top_items[item_name]['revenue'] += item.price * item.quantity
    
    top_items_list = sorted(
        [{'name': k, 'quantity': v['quantity'], 'revenue': float(v['revenue'])} 
         for k, v in top_items.items()],
        key=lambda x: x['quantity'],
        reverse=True
    )
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="top_items_7days_{today}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    title = Paragraph("Top Items Breakdown - Last 7 Days", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))
    
    data = [['Item Name', 'Quantity', 'Revenue (₹)']]
    for item in top_items_list:
        data.append([
            item['name'],
            str(item['quantity']),
            f"₹{item['revenue']:.2f}"
        ])
    
    table = Table(data, colWidths=[3*inch, 1.5*inch, 1.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(table)
    
    doc.build(elements)
    return response


@login_required
def export_top_items_7days_excel(request):
    """Export top items last 7 days as Excel"""
    today = timezone.now().date()
    user_date_joined = request.user.date_joined
    
    top_items = {}
    for i in range(6, -1, -1):
        date = today - timedelta(days=i)
        orders = Order.objects.filter(
            created_at__date=date,
            status='paid',
            created_at__gte=user_date_joined
        )
        for order in orders:
            for item in order.order_items.all():
                item_name = item.menu_item.name
                if item_name not in top_items:
                    top_items[item_name] = {'quantity': 0, 'revenue': Decimal('0')}
                top_items[item_name]['quantity'] += item.quantity
                top_items[item_name]['revenue'] += item.price * item.quantity
    
    top_items_list = sorted(
        [{'name': k, 'quantity': v['quantity'], 'revenue': float(v['revenue'])} 
         for k, v in top_items.items()],
        key=lambda x: x['quantity'],
        reverse=True
    )
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Top Items 7 Days"
    
    ws['A1'] = "Top Items Breakdown - Last 7 Days"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:C1')
    
    ws['A3'] = 'Item Name'
    ws['B3'] = 'Quantity'
    ws['C3'] = 'Revenue (₹)'
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for col in ['A3', 'B3', 'C3']:
        ws[col].fill = header_fill
        ws[col].font = header_font
    
    for i, item in enumerate(top_items_list, start=4):
        ws[f'A{i}'] = item['name']
        ws[f'B{i}'] = item['quantity']
        ws[f'C{i}'] = f"₹{item['revenue']:.2f}"
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="top_items_7days_{today}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_top_items_6months_pdf(request):
    """Export top items last 6 months as PDF"""
    today = timezone.now().date()
    user_date_joined = request.user.date_joined
    
    top_items = {}
    for i in range(5, -1, -1):
        target_date = today - timedelta(days=30*i)
        year = target_date.year
        month = target_date.month
        orders = Order.objects.filter(
            created_at__year=year,
            created_at__month=month,
            status='paid',
            created_at__gte=user_date_joined
        )
        for order in orders:
            for item in order.order_items.all():
                item_name = item.menu_item.name
                if item_name not in top_items:
                    top_items[item_name] = {'quantity': 0, 'revenue': Decimal('0')}
                top_items[item_name]['quantity'] += item.quantity
                top_items[item_name]['revenue'] += item.price * item.quantity
    
    top_items_list = sorted(
        [{'name': k, 'quantity': v['quantity'], 'revenue': float(v['revenue'])} 
         for k, v in top_items.items()],
        key=lambda x: x['quantity'],
        reverse=True
    )
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="top_items_6months_{today}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    title = Paragraph("Top Items Breakdown - Last 6 Months", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))
    
    data = [['Item Name', 'Quantity', 'Revenue (₹)']]
    for item in top_items_list:
        data.append([
            item['name'],
            str(item['quantity']),
            f"₹{item['revenue']:.2f}"
        ])
    
    table = Table(data, colWidths=[3*inch, 1.5*inch, 1.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(table)
    
    doc.build(elements)
    return response


@login_required
def export_top_items_6months_excel(request):
    """Export top items last 6 months as Excel"""
    today = timezone.now().date()
    user_date_joined = request.user.date_joined
    
    top_items = {}
    for i in range(5, -1, -1):
        target_date = today - timedelta(days=30*i)
        year = target_date.year
        month = target_date.month
        orders = Order.objects.filter(
            created_at__year=year,
            created_at__month=month,
            status='paid',
            created_at__gte=user_date_joined
        )
        for order in orders:
            for item in order.order_items.all():
                item_name = item.menu_item.name
                if item_name not in top_items:
                    top_items[item_name] = {'quantity': 0, 'revenue': Decimal('0')}
                top_items[item_name]['quantity'] += item.quantity
                top_items[item_name]['revenue'] += item.price * item.quantity
    
    top_items_list = sorted(
        [{'name': k, 'quantity': v['quantity'], 'revenue': float(v['revenue'])} 
         for k, v in top_items.items()],
        key=lambda x: x['quantity'],
        reverse=True
    )
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Top Items 6 Months"
    
    ws['A1'] = "Top Items Breakdown - Last 6 Months"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:C1')
    
    ws['A3'] = 'Item Name'
    ws['B3'] = 'Quantity'
    ws['C3'] = 'Revenue (₹)'
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for col in ['A3', 'B3', 'C3']:
        ws[col].fill = header_fill
        ws[col].font = header_font
    
    for i, item in enumerate(top_items_list, start=4):
        ws[f'A{i}'] = item['name']
        ws[f'B{i}'] = item['quantity']
        ws[f'C{i}'] = f"₹{item['revenue']:.2f}"
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="top_items_6months_{today}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_monthly_breakdown_pdf(request):
    """Export monthly breakdown as PDF"""
    today = timezone.now().date()
    user_date_joined = request.user.date_joined
    
    monthly_breakdown = []
    for i in range(11, -1, -1):
        target_date = today - timedelta(days=30*i)
        year = target_date.year
        month = target_date.month
        monthly_sales = calculate_monthly_sales(year, month, user_date_joined)
        monthly_breakdown.append({
            'year': year,
            'month': month,
            'month_str': f"{year}-{month:02d}",
            'sales': float(monthly_sales['total_sales']),
            'orders': monthly_sales['total_orders']
        })
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="monthly_breakdown_{today}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    title = Paragraph("Monthly Breakdown - Last 12 Months", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))
    
    data = [['Month', 'Sales (₹)', 'Orders']]
    for item in monthly_breakdown:
        data.append([
            item['month_str'],
            f"₹{item['sales']:.2f}",
            str(item['orders'])
        ])
    
    table = Table(data, colWidths=[2*inch, 2*inch, 2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(table)
    
    doc.build(elements)
    return response


@login_required
def export_monthly_breakdown_excel(request):
    """Export monthly breakdown as Excel"""
    today = timezone.now().date()
    user_date_joined = request.user.date_joined
    
    monthly_breakdown = []
    for i in range(11, -1, -1):
        target_date = today - timedelta(days=30*i)
        year = target_date.year
        month = target_date.month
        monthly_sales = calculate_monthly_sales(year, month, user_date_joined)
        monthly_breakdown.append({
            'year': year,
            'month': month,
            'month_str': f"{year}-{month:02d}",
            'sales': float(monthly_sales['total_sales']),
            'orders': monthly_sales['total_orders']
        })
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Breakdown"
    
    ws['A1'] = "Monthly Breakdown - Last 12 Months"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:C1')
    
    ws['A3'] = 'Month'
    ws['B3'] = 'Sales (₹)'
    ws['C3'] = 'Orders'
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for col in ['A3', 'B3', 'C3']:
        ws[col].fill = header_fill
        ws[col].font = header_font
    
    for i, item in enumerate(monthly_breakdown, start=4):
        ws[f'A{i}'] = item['month_str']
        ws[f'B{i}'] = f"₹{item['sales']:.2f}"
        ws[f'C{i}'] = item['orders']
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="monthly_breakdown_{today}.xlsx"'
    wb.save(response)
    return response
